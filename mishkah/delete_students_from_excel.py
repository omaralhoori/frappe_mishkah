# Copyright (c) 2026, Omar Alhori and contributors
# For license information, please see license.txt

"""
Delete Mishkah students (and related joining requests) whose mobiles are listed
in an Excel file (column of phone numbers).

Safe for busy sites during registration:
- one phone / student per short transaction
- retry on lock wait timeout
- small batches on the long queue with sleep between students and batches
"""

import os
import time

import frappe
from frappe.exceptions import QueryTimeoutError

BATCH_SIZE = 10
MAX_RETRIES = 5
RETRY_SLEEP_SEC = 2
BETWEEN_STUDENTS_SLEEP_SEC = 0.4
BETWEEN_BATCHES_SLEEP_SEC = 3
EMPTY_JOB_BACKOFF_SEC = 5
JOB_NAME = "mishkah_delete_students_from_excel"
CACHE_KEY = "mishkah_delete_students_from_excel_phones"
DEFAULT_EXCEL_RELATIVE = "delete_students_mishkah.xlsx"


def default_excel_path():
	"""sites/delete_students_mishkah.xlsx at the bench root."""
	return os.path.join(frappe.utils.get_bench_path(), "sites", DEFAULT_EXCEL_RELATIVE)


def normalize_phone(phone):
	"""Strip whitespace; keep digits and leading + for matching."""
	if phone is None:
		return ""
	raw = str(phone).strip()
	if not raw or raw.lower() in ("nan", "none", "numbers"):
		return ""
	# Excel may store numbers as floats (e.g. 16137002914.0)
	if isinstance(phone, float):
		raw = str(int(phone)) if phone == int(phone) else str(phone).strip()
	digits = "".join(ch for ch in raw if ch.isdigit())
	return digits


def load_phones_from_excel(file_path=None):
	"""
	Read unique phone numbers from the first column of the Excel file.
	Skips header row if the first cell is non-numeric text (e.g. NUMBERS).
	"""
	from openpyxl import load_workbook

	path = file_path or default_excel_path()
	if not os.path.exists(path):
		frappe.throw(f"Excel file not found: {path}")

	wb = load_workbook(path, read_only=True, data_only=True)
	ws = wb.active

	phones = []
	seen = set()
	first = True

	for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
		value = row[0]
		if first:
			first = False
			# Skip header like "NUMBERS"
			if value is not None and not str(value).strip().replace(".", "", 1).isdigit() and "-" not in str(value):
				normalized_header = normalize_phone(value)
				if not normalized_header:
					continue

		phone = normalize_phone(value)
		if not phone or phone in seen:
			continue
		seen.add(phone)
		phones.append(phone)

	wb.close()
	return phones


def count_matches(phones=None, file_path=None):
	"""Preview how many students / joining requests match the Excel phones."""
	phones = phones if phones is not None else load_phones_from_excel(file_path)
	if not phones:
		return {
			"phones_in_file": 0,
			"students": 0,
			"joining_requests": 0,
		}

	# Chunk IN clauses to avoid huge queries
	students = 0
	joining = 0
	chunk = 500
	for i in range(0, len(phones), chunk):
		part = phones[i : i + chunk]
		students += frappe.db.sql(
			"""
			SELECT COUNT(*) FROM `tabMishkah Student`
			WHERE student_mobile IN %(phones)s
			""",
			{"phones": tuple(part)},
		)[0][0]
		joining += frappe.db.sql(
			"""
			SELECT COUNT(*) FROM `tabMishkah Student Joining Request`
			WHERE CONCAT(IFNULL(country_code, ''), IFNULL(mobile_phone, '')) IN %(phones)s
			""",
			{"phones": tuple(part)},
		)[0][0]

	return {
		"phones_in_file": len(phones),
		"students": students,
		"joining_requests": joining,
	}


def _delete_joining_requests_for_mobile(mobile):
	frappe.db.sql(
		"""
		DELETE FROM `tabMishkah Student Joining Request`
		WHERE CONCAT(IFNULL(country_code, ''), IFNULL(mobile_phone, '')) = %(mobile)s
		""",
		{"mobile": mobile},
	)


def _delete_student_and_related(student):
	"""Delete one student and all known related rows (raw SQL, short locks)."""
	frappe.db.sql(
		"""
		DELETE FROM `tabMishkah Course Progress`
		WHERE student = %(student)s
		""",
		{"student": student},
	)

	frappe.db.sql(
		"""
		DELETE FROM `tabMishkah Student Group Student`
		WHERE student = %(student)s
		""",
		{"student": student},
	)

	frappe.db.sql(
		"""
		DELETE FROM `tabMishkah Level Complete Tool Student`
		WHERE student = %(student)s
		""",
		{"student": student},
	)

	program_enrollments = frappe.db.sql(
		"""
		SELECT name FROM `tabMishkah Program Enrollment`
		WHERE student = %(student)s
		""",
		{"student": student},
		as_dict=True,
	)

	for pe in program_enrollments:
		frappe.db.sql(
			"""
			DELETE FROM `tabMishkah Level Enrollment`
			WHERE program_enrollment = %(program_enrollment)s
			""",
			{"program_enrollment": pe.name},
		)
		frappe.db.sql(
			"""
			DELETE FROM `tabMishkah Program Enrollment`
			WHERE name = %(program_enrollment)s
			""",
			{"program_enrollment": pe.name},
		)

	frappe.db.sql(
		"""
		DELETE FROM `tabMishkah Student`
		WHERE name = %(student)s
		""",
		{"student": student},
	)


def delete_one_by_mobile(mobile):
	"""
	Delete joining request(s) for mobile, and the student if present.
	Returns dict with what was deleted. False-ish skipped on lock failure.
	"""
	for attempt in range(1, MAX_RETRIES + 1):
		try:
			student = frappe.db.get_value(
				"Mishkah Student", {"student_mobile": mobile}, "name"
			)

			_delete_joining_requests_for_mobile(mobile)

			if student:
				_delete_student_and_related(student)

			frappe.db.commit()
			return {
				"ok": True,
				"mobile": mobile,
				"student": student,
				"deleted_student": bool(student),
			}

		except QueryTimeoutError:
			frappe.db.rollback()
			if attempt >= MAX_RETRIES:
				frappe.logger("mishkah").warning(
					f"Excel student delete skipped {mobile} after {MAX_RETRIES} lock timeouts"
				)
				return {"ok": False, "mobile": mobile, "student": None, "deleted_student": False}
			time.sleep(RETRY_SLEEP_SEC * attempt)


def delete_phones_batch(phones):
	"""Delete one phone at a time with sleep to reduce DB pressure."""
	if not phones:
		return {"deleted_students": 0, "deleted_joining_only": 0, "skipped": 0, "not_found": 0}

	deleted_students = 0
	deleted_joining_only = 0
	skipped = 0
	not_found = 0

	for mobile in phones:
		# Fast existence check without holding long transactions for misses
		has_student = frappe.db.exists("Mishkah Student", {"student_mobile": mobile})
		has_joining = frappe.db.sql(
			"""
			SELECT 1 FROM `tabMishkah Student Joining Request`
			WHERE CONCAT(IFNULL(country_code, ''), IFNULL(mobile_phone, '')) = %(mobile)s
			LIMIT 1
			""",
			{"mobile": mobile},
		)

		if not has_student and not has_joining:
			not_found += 1
			time.sleep(BETWEEN_STUDENTS_SLEEP_SEC)
			continue

		result = delete_one_by_mobile(mobile)
		if not result or not result.get("ok"):
			skipped += 1
		elif result.get("deleted_student"):
			deleted_students += 1
		else:
			deleted_joining_only += 1

		time.sleep(BETWEEN_STUDENTS_SLEEP_SEC)

	return {
		"deleted_students": deleted_students,
		"deleted_joining_only": deleted_joining_only,
		"skipped": skipped,
		"not_found": not_found,
	}


def delete_students_from_excel_job(offset=0, batch_size=BATCH_SIZE, empty_streak=0):
	"""Process one batch from cached phone list, then enqueue the next."""
	offset = int(offset or 0)
	batch_size = int(batch_size or BATCH_SIZE)
	empty_streak = int(empty_streak or 0)

	phones = frappe.cache().get_value(CACHE_KEY)
	if not phones:
		frappe.logger("mishkah").error(
			"Excel student delete: phone list missing from cache; re-enqueue from Desk"
		)
		return {"done": True, "error": "cache_empty"}

	batch = phones[offset : offset + batch_size]
	if not batch:
		frappe.cache().delete_value(CACHE_KEY)
		frappe.logger("mishkah").info("Excel student delete finished: no phones left")
		return {"done": True, "deleted_students": 0, "remaining": 0}

	result = delete_phones_batch(batch)
	next_offset = offset + batch_size
	remaining = max(0, len(phones) - next_offset)

	frappe.logger("mishkah").info(
		f"Excel student delete: students={result['deleted_students']}, "
		f"joining_only={result['deleted_joining_only']}, skipped={result['skipped']}, "
		f"not_found={result['not_found']}, remaining_phones~{remaining}"
	)

	if remaining == 0:
		frappe.cache().delete_value(CACHE_KEY)
		return {**result, "remaining": 0, "done": True}

	worked = result["deleted_students"] + result["deleted_joining_only"]
	if worked == 0 and result["skipped"] > 0:
		empty_streak += 1
		if empty_streak >= 5:
			frappe.logger("mishkah").error(
				"Excel student delete stopped: repeated lock timeouts. Re-run later."
			)
			return {**result, "remaining": remaining, "done": False, "stopped": True}
		time.sleep(EMPTY_JOB_BACKOFF_SEC * empty_streak)
	else:
		empty_streak = 0
		time.sleep(BETWEEN_BATCHES_SLEEP_SEC)

	frappe.enqueue(
		"mishkah.delete_students_from_excel.delete_students_from_excel_job",
		queue="long",
		timeout=2000,
		offset=next_offset,
		batch_size=batch_size,
		empty_streak=empty_streak,
		job_name=JOB_NAME,
		enqueue_after_commit=True,
	)

	return {**result, "remaining": remaining, "done": False, "next_offset": next_offset}


@frappe.whitelist()
def preview_delete_students_from_excel(file_path=None):
	"""
	Preview matches without deleting.
	Console: mishkah.delete_students_from_excel.preview_delete_students_from_excel()
	"""
	frappe.only_for("System Manager")
	phones = load_phones_from_excel(file_path)
	stats = count_matches(phones=phones)
	stats["file_path"] = file_path or default_excel_path()
	return stats


@frappe.whitelist()
def enqueue_delete_students_from_excel(file_path=None, batch_size=BATCH_SIZE):
	"""
	Start background delete from Excel. Safe during registration (gentle pacing).

	Preview first:
	  mishkah.delete_students_from_excel.preview_delete_students_from_excel()

	Then start:
	  mishkah.delete_students_from_excel.enqueue_delete_students_from_excel()
	"""
	frappe.only_for("System Manager")
	batch_size = int(batch_size or BATCH_SIZE)

	phones = load_phones_from_excel(file_path)
	if not phones:
		return {"queued": False, "message": "No phone numbers found in Excel"}

	stats = count_matches(phones=phones)

	# Keep list in cache for the duration of the job chain (~1 day)
	frappe.cache().set_value(CACHE_KEY, phones, expires_in_sec=86400)

	frappe.enqueue(
		"mishkah.delete_students_from_excel.delete_students_from_excel_job",
		queue="long",
		timeout=2000,
		offset=0,
		batch_size=batch_size,
		empty_streak=0,
		job_name=JOB_NAME,
	)

	return {
		"queued": True,
		"phones_in_file": stats["phones_in_file"],
		"students_matched": stats["students"],
		"joining_requests_matched": stats["joining_requests"],
		"batch_size": batch_size,
		"file_path": file_path or default_excel_path(),
		"message": (
			f"Queued gentle delete for {stats['phones_in_file']} phones "
			f"(~{stats['students']} students, ~{stats['joining_requests']} joining requests); "
			f"batch size {batch_size}"
		),
	}
